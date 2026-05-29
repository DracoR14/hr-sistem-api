"""
EVOS v2.0 — HR Management System
PC Desktop aplikacija
- Edit korisnika i licenci
- Logo firme + EVOS branding
- Grafikoni u PDF i Excel
- Poboljšan dizajn
- Višejezičnost
"""

import sys, os, json, requests, base64, io
from datetime import date, datetime
from pathlib import Path

from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QStackedWidget, QLabel, QPushButton, QTableWidget, QTableWidgetItem,
    QDialog, QFormLayout, QLineEdit, QDateEdit, QComboBox, QTextEdit,
    QMessageBox, QHeaderView, QFrame, QTabWidget, QFileDialog, QCheckBox,
    QListWidget, QListWidgetItem, QInputDialog, QAbstractItemView,
    QScrollArea, QSizePolicy, QSpinBox
)
from PyQt6.QtCore import Qt, QDate, QTimer, QThread, pyqtSignal, QSize
from PyQt6.QtGui import (
    QFont, QColor, QPainter, QBrush, QCursor, QPixmap, QImage, QPalette
)
from PyQt6.QtCharts import (
    QChart, QChartView, QBarSeries, QBarSet, QBarCategoryAxis,
    QValueAxis, QPieSeries, QPieSlice
)
from PyQt6.QtPrintSupport import QPrinter, QPrintPreviewDialog

try:
    import openpyxl
    from openpyxl.styles import Font as XFont, PatternFill, Alignment, Border, Side
    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.chart.series import DataPoint
    import openpyxl.utils
    HAS_XL = True
except ImportError:
    HAS_XL = False

try:
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors as rlc
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
        Paragraph, Spacer, Image as RLImage, HRFlowable)
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.graphics.shapes import Drawing, Wedge, Rect, String
    from reportlab.graphics.charts.piecharts import Pie
    from reportlab.graphics.charts.barcharts import VerticalBarChart
    from reportlab.graphics import renderPDF
    HAS_PDF = True
except ImportError:
    HAS_PDF = False

# ── Config ─────────────────────────────────────────────────────────────────────
API_URL    = "https://hr-sistem-api-production.up.railway.app"
CONFIG_DIR = Path(os.path.expanduser("~")) / "EVOS"
CONFIG_DIR.mkdir(exist_ok=True)
CONFIG_FILE = CONFIG_DIR / "config.json"
LOGO_FILE   = CONFIG_DIR / "logo.png"

# ── Translations ───────────────────────────────────────────────────────────────
T = {
    "bs": {
        "app_name":"EVOS","login_title":"Dobrodošli u EVOS",
        "login_sub":"Prijavite se da nastavite",
        "login_user":"Korisničko ime","login_pass":"Lozinka",
        "login_btn":"Prijava","login_err":"Pogrešno korisničko ime ili lozinka",
        "logout":"Odjava",
        "nav_dashboard":"Dashboard","nav_uposlenici":"Uposlenici",
        "nav_bolovanja":"Bolovanja","nav_odsustva":"Odsustva",
        "nav_statistike":"Statistike","nav_postavke":"Postavke","nav_licence":"Licence",
        "up_title":"Uposlenici","up_novi":"+ Novi uposlenik",
        "bol_title":"Bolovanja","bol_novo":"+ Novo bolovanje",
        "ods_title":"Odsustva","ods_novo":"+ Novo odsustvo",
        "stat_title":"Statistike","post_title":"Postavke","lic_title":"Licence",
        "lic_nova":"+ Nova licenca","lic_edit":"Uredi licencu",
        "btn_save":"Spremi","btn_cancel":"Odustani","btn_del":"Obriši",
        "btn_edit":"Uredi","btn_excel":"📊 Excel","btn_pdf":"📄 PDF","btn_print":"🖨 Print",
        "lbl_search":"Pretraži...","lbl_dept":"Služba","lbl_status":"Status",
        "lbl_pos":"Pozicija","lbl_hire":"Datum zaposlenja",
        "lbl_from":"Datum od","lbl_to":"Datum do","lbl_days":"Dana",
        "lbl_code":"Šifra","lbl_reason":"Razlog","lbl_note":"Napomena",
        "lbl_type":"Tip","lbl_firm":"Firma","lbl_plan":"Plan",
        "lbl_expires":"Ističe","lbl_max_emp":"Max uposlenika",
        "lbl_key":"Licencni ključ","lbl_lang":"Jezik","lbl_emp":"Uposlenik",
        "lbl_logo":"Logo firme","lbl_upload_logo":"Učitaj logo",
        "status_active":"Aktivan","status_inactive":"Neaktivan","status_probation":"Na probnom radu",
        "ods_pending":"Na čekanju","ods_approved":"Odobreno","ods_rejected":"Odbijeno",
        "ods_approve":"Odobri","ods_reject":"Odbij",
        "err_required":"Ovo polje je obavezno!","err_dates":"'Do' mora biti >= 'Od'!",
        "del_confirm":"Jeste li sigurni da želite obrisati?",
        "conn_err":"Greška konekcije. Provjerite internet vezu.",
        "lang_bs":"Bosanski","lang_hr":"Hrvatski","lang_sr":"Srpski","lang_en":"English",
        "report_title":"Izvještaj","report_date":"Datum izvještaja",
        "report_generated":"Generirao","report_total":"Ukupno dana",
        "report_by_emp":"Dana bolovanja po uposleniku",
        "report_by_month":"Trend po mjesecu","report_by_dept":"Po službi",
    },
    "hr": {
        "app_name":"EVOS","login_title":"Dobrodošli u EVOS",
        "login_sub":"Prijavite se za nastavak",
        "login_user":"Korisničko ime","login_pass":"Lozinka",
        "login_btn":"Prijava","login_err":"Pogrešno korisničko ime ili lozinka",
        "logout":"Odjava",
        "nav_dashboard":"Pregled","nav_uposlenici":"Zaposlenici",
        "nav_bolovanja":"Bolovanja","nav_odsustva":"Odsustva",
        "nav_statistike":"Statistike","nav_postavke":"Postavke","nav_licence":"Licence",
        "up_title":"Zaposlenici","up_novi":"+ Novi zaposlenik",
        "bol_title":"Bolovanja","bol_novo":"+ Novo bolovanje",
        "ods_title":"Odsustva","ods_novo":"+ Novo odsustvo",
        "stat_title":"Statistike","post_title":"Postavke","lic_title":"Licence",
        "lic_nova":"+ Nova licenca","lic_edit":"Uredi licencu",
        "btn_save":"Spremi","btn_cancel":"Odustani","btn_del":"Obriši",
        "btn_edit":"Uredi","btn_excel":"📊 Excel","btn_pdf":"📄 PDF","btn_print":"🖨 Ispis",
        "lbl_search":"Pretraži...","lbl_dept":"Odjel","lbl_status":"Status",
        "lbl_pos":"Radno mjesto","lbl_hire":"Datum zaposlenja",
        "lbl_from":"Datum od","lbl_to":"Datum do","lbl_days":"Dana",
        "lbl_code":"Šifra","lbl_reason":"Razlog","lbl_note":"Napomena",
        "lbl_type":"Vrsta","lbl_firm":"Tvrtka","lbl_plan":"Plan",
        "lbl_expires":"Ističe","lbl_max_emp":"Maks. zaposlenika",
        "lbl_key":"Licencni ključ","lbl_lang":"Jezik","lbl_emp":"Zaposlenik",
        "lbl_logo":"Logo tvrtke","lbl_upload_logo":"Učitaj logo",
        "status_active":"Aktivan","status_inactive":"Neaktivan","status_probation":"Na probnom radu",
        "ods_pending":"Na čekanju","ods_approved":"Odobreno","ods_rejected":"Odbijeno",
        "ods_approve":"Odobri","ods_reject":"Odbij",
        "err_required":"Ovo polje je obavezno!","err_dates":"'Do' mora biti >= 'Od'!",
        "del_confirm":"Jeste li sigurni da želite obrisati?",
        "conn_err":"Greška veze. Provjerite internetsku vezu.",
        "lang_bs":"Bosanski","lang_hr":"Hrvatski","lang_sr":"Srpski","lang_en":"English",
        "report_title":"Izvještaj","report_date":"Datum izvještaja",
        "report_generated":"Generirao","report_total":"Ukupno dana",
        "report_by_emp":"Dana bolovanja po zaposleniku",
        "report_by_month":"Trend po mjesecu","report_by_dept":"Po odjelu",
    },
    "sr": {
        "app_name":"EVOS","login_title":"Dobrodošli u EVOS",
        "login_sub":"Prijavite se da nastavite",
        "login_user":"Korisničko ime","login_pass":"Lozinka",
        "login_btn":"Prijava","login_err":"Pogrešno korisničko ime ili lozinka",
        "logout":"Odjava",
        "nav_dashboard":"Pregled","nav_uposlenici":"Zaposleni",
        "nav_bolovanja":"Bolovanja","nav_odsustva":"Odsustva",
        "nav_statistike":"Statistike","nav_postavke":"Podešavanja","nav_licence":"Licence",
        "up_title":"Zaposleni","up_novi":"+ Novi zaposleni",
        "bol_title":"Bolovanja","bol_novo":"+ Novo bolovanje",
        "ods_title":"Odsustva","ods_novo":"+ Novo odsustvo",
        "stat_title":"Statistike","post_title":"Podešavanja","lic_title":"Licence",
        "lic_nova":"+ Nova licenca","lic_edit":"Uredi licencu",
        "btn_save":"Sačuvaj","btn_cancel":"Odustani","btn_del":"Obriši",
        "btn_edit":"Uredi","btn_excel":"📊 Excel","btn_pdf":"📄 PDF","btn_print":"🖨 Štampa",
        "lbl_search":"Pretraži...","lbl_dept":"Služba","lbl_status":"Status",
        "lbl_pos":"Pozicija","lbl_hire":"Datum zaposlenja",
        "lbl_from":"Datum od","lbl_to":"Datum do","lbl_days":"Dana",
        "lbl_code":"Šifra","lbl_reason":"Razlog","lbl_note":"Napomena",
        "lbl_type":"Tip","lbl_firm":"Firma","lbl_plan":"Plan",
        "lbl_expires":"Ističe","lbl_max_emp":"Maks. zaposlenih",
        "lbl_key":"Licencni ključ","lbl_lang":"Jezik","lbl_emp":"Zaposleni",
        "lbl_logo":"Logo firme","lbl_upload_logo":"Učitaj logo",
        "status_active":"Aktivan","status_inactive":"Neaktivan","status_probation":"Na probnom radu",
        "ods_pending":"Na čekanju","ods_approved":"Odobreno","ods_rejected":"Odbijeno",
        "ods_approve":"Odobri","ods_reject":"Odbij",
        "err_required":"Ovo polje je obavezno!","err_dates":"'Do' mora biti >= 'Od'!",
        "del_confirm":"Da li ste sigurni da želite obrisati?",
        "conn_err":"Greška konekcije. Proverite internet vezu.",
        "lang_bs":"Bosanski","lang_hr":"Hrvatski","lang_sr":"Srpski","lang_en":"English",
        "report_title":"Izveštaj","report_date":"Datum izveštaja",
        "report_generated":"Generisao","report_total":"Ukupno dana",
        "report_by_emp":"Dana bolovanja po zaposlenom",
        "report_by_month":"Trend po mesecu","report_by_dept":"Po službi",
    },
    "en": {
        "app_name":"EVOS","login_title":"Welcome to EVOS",
        "login_sub":"Sign in to continue",
        "login_user":"Username","login_pass":"Password",
        "login_btn":"Sign in","login_err":"Incorrect username or password",
        "logout":"Sign out",
        "nav_dashboard":"Dashboard","nav_uposlenici":"Employees",
        "nav_bolovanja":"Sick Leave","nav_odsustva":"Absences",
        "nav_statistike":"Statistics","nav_postavke":"Settings","nav_licence":"Licenses",
        "up_title":"Employees","up_novi":"+ New employee",
        "bol_title":"Sick Leave","bol_novo":"+ New sick leave",
        "ods_title":"Absences","ods_novo":"+ New absence",
        "stat_title":"Statistics","post_title":"Settings","lic_title":"Licenses",
        "lic_nova":"+ New license","lic_edit":"Edit license",
        "btn_save":"Save","btn_cancel":"Cancel","btn_del":"Delete",
        "btn_edit":"Edit","btn_excel":"📊 Excel","btn_pdf":"📄 PDF","btn_print":"🖨 Print",
        "lbl_search":"Search...","lbl_dept":"Department","lbl_status":"Status",
        "lbl_pos":"Position","lbl_hire":"Hire date",
        "lbl_from":"From date","lbl_to":"To date","lbl_days":"Days",
        "lbl_code":"Code","lbl_reason":"Reason","lbl_note":"Note",
        "lbl_type":"Type","lbl_firm":"Company","lbl_plan":"Plan",
        "lbl_expires":"Expires","lbl_max_emp":"Max employees",
        "lbl_key":"License key","lbl_lang":"Language","lbl_emp":"Employee",
        "lbl_logo":"Company logo","lbl_upload_logo":"Upload logo",
        "status_active":"Active","status_inactive":"Inactive","status_probation":"Probation",
        "ods_pending":"Pending","ods_approved":"Approved","ods_rejected":"Rejected",
        "ods_approve":"Approve","ods_reject":"Reject",
        "err_required":"This field is required!","err_dates":"'To' must be >= 'From'!",
        "del_confirm":"Are you sure you want to delete?",
        "conn_err":"Connection error. Check your internet connection.",
        "lang_bs":"Bosanski","lang_hr":"Hrvatski","lang_sr":"Srpski","lang_en":"English",
        "report_title":"Report","report_date":"Report date",
        "report_generated":"Generated by","report_total":"Total days",
        "report_by_emp":"Sick days by employee",
        "report_by_month":"Monthly trend","report_by_dept":"By department",
    },
}

class AppState:
    token = ""; user = {}; lang = "bs"; logo_path = ""
    @classmethod
    def tr(cls, key): return T.get(cls.lang, T["bs"]).get(key, key)
    @classmethod
    def save(cls):
        with open(CONFIG_FILE,"w") as f:
            json.dump({"lang":cls.lang,"token":cls.token,"user":cls.user,"logo":cls.logo_path},f)
    @classmethod
    def load(cls):
        if CONFIG_FILE.exists():
            try:
                d=json.loads(CONFIG_FILE.read_text())
                cls.lang=d.get("lang","bs"); cls.token=d.get("token","")
                cls.user=d.get("user",{}); cls.logo_path=d.get("logo","")
            except: pass

S = AppState

# ── API ────────────────────────────────────────────────────────────────────────
class API:
    @staticmethod
    def _h(): return {"Authorization":f"Bearer {S.token}","Content-Type":"application/json"}
    @staticmethod
    def get(path, params=None):
        try:
            r=requests.get(f"{API_URL}{path}",headers=API._h(),params=params,timeout=15)
            r.raise_for_status(); return r.json()
        except requests.exceptions.ConnectionError: raise Exception(S.tr("conn_err"))
        except Exception as e: raise Exception(str(e))
    @staticmethod
    def post(path, data):
        try:
            r=requests.post(f"{API_URL}{path}",headers=API._h(),json=data,timeout=15)
            r.raise_for_status(); return r.json()
        except requests.exceptions.ConnectionError: raise Exception(S.tr("conn_err"))
        except Exception as e: raise Exception(str(e))
    @staticmethod
    def put(path, data):
        try:
            r=requests.put(f"{API_URL}{path}",headers=API._h(),json=data,timeout=15)
            r.raise_for_status(); return r.json()
        except requests.exceptions.ConnectionError: raise Exception(S.tr("conn_err"))
        except Exception as e: raise Exception(str(e))
    @staticmethod
    def patch(path, data):
        try:
            r=requests.patch(f"{API_URL}{path}",headers=API._h(),json=data,timeout=15)
            r.raise_for_status(); return r.json()
        except: raise Exception(S.tr("conn_err"))
    @staticmethod
    def delete(path):
        try:
            r=requests.delete(f"{API_URL}{path}",headers=API._h(),timeout=15)
            r.raise_for_status(); return r.json()
        except requests.exceptions.ConnectionError: raise Exception(S.tr("conn_err"))
        except Exception as e: raise Exception(str(e))
    @staticmethod
    def login(username, password):
        try:
            r=requests.post(f"{API_URL}/auth/login",data={"username":username,"password":password},timeout=15)
            if r.status_code==401: return None, S.tr("login_err")
            r.raise_for_status(); return r.json(), None
        except requests.exceptions.ConnectionError: return None, S.tr("conn_err")
        except Exception as e: return None, str(e)

# ── Palette ────────────────────────────────────────────────────────────────────
P = {
    "bg":"#0F1117","surface":"#1A1D2E","card":"#222539",
    "border":"#2E3250","accent":"#4F6EF7","accent2":"#7C3AED",
    "green":"#22C55E","orange":"#F59E0B","red":"#EF4444",
    "text":"#F1F5F9","muted":"#94A3B8","hover":"#2E3355",
}

STYLE = f"""
QMainWindow,QWidget{{background:{P['bg']};color:{P['text']};font-family:'Segoe UI';font-size:13px;}}
QDialog{{background:{P['surface']};color:{P['text']};}}
QLabel{{color:{P['text']};}}
QLineEdit,QTextEdit,QDateEdit,QComboBox,QSpinBox{{
    background:{P['card']};border:1.5px solid {P['border']};border-radius:8px;
    padding:8px 12px;color:{P['text']};font-size:13px;}}
QLineEdit:focus,QTextEdit:focus,QDateEdit:focus,QComboBox:focus,QSpinBox:focus{{border-color:{P['accent']};}}
QComboBox::drop-down{{border:none;width:28px;}}
QComboBox QAbstractItemView{{background:{P['card']};border:1px solid {P['border']};
    selection-background-color:{P['accent']};color:{P['text']};}}
QTableWidget{{background:{P['surface']};border:none;gridline-color:{P['border']};color:{P['text']};}}
QTableWidget::item{{padding:6px 12px;border:none;}}
QTableWidget::item:selected{{background:{P['accent']}33;color:{P['text']};}}
QTableWidget::item:alternate{{background:{P['card']};}}
QHeaderView::section{{background:{P['card']};color:{P['muted']};font-size:11px;
    font-weight:600;letter-spacing:1px;padding:10px 12px;border:none;
    border-bottom:1.5px solid {P['border']};}}
QScrollBar:vertical{{background:{P['surface']};width:8px;border-radius:4px;}}
QScrollBar::handle:vertical{{background:{P['border']};border-radius:4px;min-height:30px;}}
QScrollBar::add-line:vertical,QScrollBar::sub-line:vertical{{height:0;}}
QTabWidget::pane{{border:1.5px solid {P['border']};border-radius:10px;background:{P['surface']};}}
QTabBar::tab{{background:{P['card']};color:{P['muted']};padding:10px 22px;
    border-radius:8px 8px 0 0;font-size:13px;font-weight:600;margin-right:3px;}}
QTabBar::tab:selected{{background:{P['accent']};color:white;}}
QCheckBox{{color:{P['text']};font-size:13px;spacing:8px;}}
QCheckBox::indicator{{width:18px;height:18px;border:2px solid {P['border']};border-radius:4px;background:{P['card']};}}
QCheckBox::indicator:checked{{background:{P['accent']};border-color:{P['accent']};}}
QListWidget{{background:{P['card']};border:1.5px solid {P['border']};border-radius:8px;color:{P['text']};}}
QListWidget::item{{padding:8px 12px;border-radius:6px;}}
QListWidget::item:selected{{background:{P['accent']}33;color:{P['accent']};}}
"""

# ── UI Helpers ────────────────────────────────────────────────────────────────
def card():
    w=QFrame()
    w.setStyleSheet(f"QFrame{{background:{P['surface']};border:1.5px solid {P['border']};border-radius:14px;}}")
    return w

def btn_primary(t):
    b=QPushButton(t)
    b.setStyleSheet(f"QPushButton{{background:{P['accent']};color:white;border:none;border-radius:8px;padding:9px 20px;font-size:13px;font-weight:600;}}QPushButton:hover{{background:#6B84FF;}}QPushButton:pressed{{background:#3A54D4;}}")
    b.setCursor(QCursor(Qt.CursorShape.PointingHandCursor)); return b

def btn_ghost(t):
    b=QPushButton(t)
    b.setStyleSheet(f"QPushButton{{background:transparent;color:{P['muted']};border:1.5px solid {P['border']};border-radius:8px;padding:9px 16px;font-size:13px;}}QPushButton:hover{{color:{P['text']};border-color:{P['accent']};}}")
    b.setCursor(QCursor(Qt.CursorShape.PointingHandCursor)); return b

def btn_danger(t):
    b=QPushButton(t)
    b.setStyleSheet(f"QPushButton{{background:{P['red']}22;color:{P['red']};border:1.5px solid {P['red']}55;border-radius:8px;padding:9px 16px;font-size:13px;font-weight:600;}}QPushButton:hover{{background:{P['red']}44;}}")
    b.setCursor(QCursor(Qt.CursorShape.PointingHandCursor)); return b

def btn_icon(icon, color=None):
    c=color or P['muted']; b=QPushButton(icon); b.setFixedSize(34,32)
    b.setStyleSheet(f"QPushButton{{background:{P['card']};color:{c};border:1px solid {P['border']};border-radius:6px;font-size:14px;}}QPushButton:hover{{background:{P['hover']};border-color:{P['accent']};}}")
    b.setCursor(QCursor(Qt.CursorShape.PointingHandCursor)); return b

def stat_card(title, value, color, sub=""):
    f=card(); f.setFixedHeight(110)
    lay=QVBoxLayout(f); lay.setContentsMargins(20,16,20,16); lay.setSpacing(4)
    t=QLabel(title); t.setStyleSheet(f"color:{P['muted']};font-size:11px;font-weight:600;letter-spacing:1px;")
    v=QLabel(str(value)); v.setStyleSheet(f"color:{color};font-size:30px;font-weight:700;")
    lay.addWidget(t); lay.addWidget(v)
    if sub:
        s=QLabel(sub); s.setStyleSheet(f"color:{P['muted']};font-size:11px;"); lay.addWidget(s)
    return f

def make_table(cols):
    t=QTableWidget(); t.setColumnCount(len(cols))
    t.setHorizontalHeaderLabels(cols)
    t.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
    t.verticalHeader().setVisible(False); t.setAlternatingRowColors(True)
    t.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
    t.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
    t.verticalHeader().setDefaultSectionSize(46); return t

def titem(v, center=False):
    it=QTableWidgetItem(str(v) if v is not None else "")
    it.setTextAlignment(Qt.AlignmentFlag.AlignVCenter|(Qt.AlignmentFlag.AlignCenter if center else Qt.AlignmentFlag.AlignLeft))
    return it

def api_error(parent, e): QMessageBox.warning(parent,"Greška",str(e))

QLabel.also = lambda self, fn: (fn(self), self)[1]

# ── EXPORT SA GRAFIKONIMA ──────────────────────────────────────────────────────
def export_excel_with_charts(parent, headers, rows, stats, title="Izvjestaj"):
    if not HAS_XL:
        QMessageBox.warning(parent,"Greška","pip install openpyxl"); return
    path,_=QFileDialog.getSaveFileName(parent,"Spremi Excel",str(CONFIG_DIR/f"{title}.xlsx"),"Excel (*.xlsx)")
    if not path: return
    wb=openpyxl.Workbook()

    # Sheet 1 - Podaci
    ws=wb.active; ws.title="Podaci"
    thin=Side(style='thin',color='B0C4DE'); brd=Border(left=thin,right=thin,top=thin,bottom=thin)
    hf=PatternFill("solid",start_color="2E75B6")
    for ci,h in enumerate(headers,1):
        c=ws.cell(1,ci,h); c.font=XFont(bold=True,color="FFFFFF",size=11)
        c.fill=hf; c.alignment=Alignment(horizontal="center",vertical="center"); c.border=brd
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width=max(14,len(str(h))+4)
    af=PatternFill("solid",start_color="D6E4F0")
    for ri,row in enumerate(rows,2):
        for ci,v in enumerate(row,1):
            c=ws.cell(ri,ci,v); c.border=brd; c.alignment=Alignment(vertical="center")
            if ri%2==0: c.fill=af

    # Sheet 2 - Grafikoni
    wc=wb.create_sheet("Grafikoni")
    wc["A1"]="Analiza bolovanja"; wc["A1"].font=XFont(bold=True,size=14,color="2E75B6")

    # Pie chart po uposlenicima
    by_emp=stats.get("by_emp",[])
    if by_emp:
        wc["A3"]="Uposlenik"; wc["B3"]="Dana"
        wc["A3"].font=XFont(bold=True); wc["B3"].font=XFont(bold=True)
        for i,row in enumerate(by_emp[:10],4):
            wc.cell(i,1,row.get("name","")[:20])
            wc.cell(i,2,row.get("dana",0))

        pie=PieChart()
        pie.title="Dana bolovanja po uposleniku"
        labels=Reference(wc,min_col=1,min_row=4,max_row=3+min(len(by_emp),10))
        data=Reference(wc,min_col=2,min_row=3,max_row=3+min(len(by_emp),10))
        pie.add_data(data,titles_from_data=True)
        pie.set_categories(labels)
        pie.style=10; pie.width=18; pie.height=14
        wc.add_chart(pie,"D3")

    # Bar chart po mjesecima
    by_month=stats.get("by_month",[])
    MJ=["Jan","Feb","Mar","Apr","Maj","Jun","Jul","Aug","Sep","Okt","Nov","Dec"]
    if by_month:
        start_row=20
        wc.cell(start_row,1,"Mjesec"); wc.cell(start_row,2,"Dana")
        mdict={r.get("mj",""):r.get("dana",0) for r in by_month}
        for i,m in enumerate(MJ,start_row+1):
            wc.cell(i,1,m); wc.cell(i,2,mdict.get(f"{MJ.index(m)+1:02d}",0))
        bar=BarChart(); bar.type="col"; bar.title="Trend bolovanja po mjesecu"
        bar.y_axis.title="Dana"; bar.x_axis.title="Mjesec"
        data=Reference(wc,min_col=2,min_row=start_row,max_row=start_row+12)
        cats=Reference(wc,min_col=1,min_row=start_row+1,max_row=start_row+12)
        bar.add_data(data,titles_from_data=True); bar.set_categories(cats)
        bar.style=10; bar.width=18; bar.height=12
        wc.add_chart(bar,"D20")

    wb.save(path)
    QMessageBox.information(parent,"Uspješno",f"Excel sa grafikonima sačuvan:\n{path}")

def export_pdf_with_charts(parent, headers, rows, stats, title="Izvjestaj", firma_naziv=""):
    if not HAS_PDF:
        QMessageBox.warning(parent,"Greška","pip install reportlab"); return
    path,_=QFileDialog.getSaveFileName(parent,"Spremi PDF",str(CONFIG_DIR/f"{title}.pdf"),"PDF (*.pdf)")
    if not path: return

    doc=SimpleDocTemplate(path,pagesize=A4,topMargin=1.5*cm,bottomMargin=1.5*cm,
                          leftMargin=1.5*cm,rightMargin=1.5*cm)
    styles=getSampleStyleSheet()
    title_style=ParagraphStyle('Title',fontName='Helvetica-Bold',fontSize=18,
                               textColor=rlc.HexColor('#2E75B6'),spaceAfter=6)
    sub_style=ParagraphStyle('Sub',fontName='Helvetica',fontSize=10,
                             textColor=rlc.HexColor('#94A3B8'),spaceAfter=12)
    section_style=ParagraphStyle('Section',fontName='Helvetica-Bold',fontSize=13,
                                 textColor=rlc.HexColor('#1A1D2E'),spaceAfter=8,spaceBefore=16)
    elems=[]

    # Header sa logom
    header_data=[[]]
    if S.logo_path and os.path.exists(S.logo_path):
        try:
            img=RLImage(S.logo_path,width=3*cm,height=2*cm)
            header_data=[[img, Paragraph(f"<b>{firma_naziv or 'EVOS'}</b>", title_style)]]
        except:
            header_data=[[Paragraph(f"<b>{firma_naziv or 'EVOS'}</b>",title_style)]]
    else:
        header_data=[[Paragraph(f"<b>EVOS</b>",title_style)]]

    if len(header_data[0])>1:
        ht=Table(header_data,colWidths=[4*cm,None])
        ht.setStyle(TableStyle([('VALIGN',(0,0),(-1,-1),'MIDDLE'),('ALIGN',(0,0),(0,0),'LEFT')]))
        elems.append(ht)
    else:
        elems.append(Paragraph("EVOS — HR Management System",title_style))

    elems.append(Paragraph(f"{title}",title_style))
    elems.append(Paragraph(f"{S.tr('report_date')}: {date.today().strftime('%d.%m.%Y.')}  |  {S.tr('report_generated')}: {S.user.get('ime','')}",sub_style))
    elems.append(HRFlowable(width="100%",thickness=1,color=rlc.HexColor('#2E75B6'),spaceAfter=12))

    # Tabela podataka
    elems.append(Paragraph(S.tr("report_title"),section_style))
    data=[headers]+[[str(v) if v is not None else "" for v in r] for r in rows]
    cw=(A4[0]-3*cm)/len(headers)
    tbl=Table(data,colWidths=[cw]*len(headers))
    tbl.setStyle(TableStyle([
        ('BACKGROUND',(0,0),(-1,0),rlc.HexColor('#2E75B6')),
        ('TEXTCOLOR',(0,0),(-1,0),rlc.white),
        ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),('FONTSIZE',(0,0),(-1,0),9),
        ('ALIGN',(0,0),(-1,-1),'CENTER'),('VALIGN',(0,0),(-1,-1),'MIDDLE'),
        ('ROWBACKGROUNDS',(0,1),(-1,-1),[rlc.white,rlc.HexColor('#EEF2FF')]),
        ('GRID',(0,0),(-1,-1),0.5,rlc.HexColor('#B0C4DE')),
        ('FONTSIZE',(0,1),(-1,-1),8),('TOPPADDING',(0,0),(-1,-1),5),('BOTTOMPADDING',(0,0),(-1,-1),5),
    ]))
    elems.append(tbl)

    # Pie chart - bolovanja po uposleniku
    by_emp=stats.get("by_emp",[])
    if by_emp:
        elems.append(Paragraph(S.tr("report_by_emp"),section_style))
        COLORS_PIE=["#4F6EF7","#7C3AED","#22C55E","#F59E0B","#EF4444","#06B6D4","#EC4899","#84CC16","#F97316","#14B8A6"]
        d=Drawing(400,200)
        pie=Pie()
        pie.x=50; pie.y=20; pie.width=160; pie.height=160
        vals=[r.get("dana",0) for r in by_emp[:8] if r.get("dana",0)>0]
        lbls=[f"{r.get('name','')[:12]} ({r.get('dana',0)}d)" for r in by_emp[:8] if r.get("dana",0)>0]
        if vals:
            pie.data=vals; pie.labels=lbls
            pie.sideLabels=True
            for i,c in enumerate(COLORS_PIE[:len(vals)]):
                pie.slices[i].fillColor=rlc.HexColor(c)
                pie.slices[i].strokeColor=rlc.white
                pie.slices[i].strokeWidth=1
            d.add(pie)
            elems.append(d)

    # Bar chart - trend po mjesecu
    by_month=stats.get("by_month",[])
    if by_month:
        elems.append(Paragraph(S.tr("report_by_month"),section_style))
        MJ=["Jan","Feb","Mar","Apr","Maj","Jun","Jul","Aug","Sep","Okt","Nov","Dec"]
        mdict={r.get("mj",""):r.get("dana",0) for r in by_month}
        month_vals=[mdict.get(f"{i+1:02d}",0) for i in range(12)]
        if any(v>0 for v in month_vals):
            d2=Drawing(500,150)
            bc=VerticalBarChart()
            bc.x=50; bc.y=10; bc.height=120; bc.width=420
            bc.data=[month_vals]
            bc.categoryAxis.categoryNames=MJ
            bc.bars[0].fillColor=rlc.HexColor('#4F6EF7')
            bc.bars[0].strokeColor=rlc.HexColor('#3A54D4')
            bc.valueAxis.valueMin=0
            bc.categoryAxis.labels.angle=30
            bc.categoryAxis.labels.fontSize=8
            bc.valueAxis.labels.fontSize=8
            d2.add(bc)
            elems.append(d2)

    doc.build(elems)
    QMessageBox.information(parent,"Uspješno",f"PDF sa grafikonima sačuvan:\n{path}")

def do_print(parent, headers, rows, title="Izvjestaj"):
    printer=QPrinter(QPrinter.PrinterMode.HighResolution)
    dlg=QPrintPreviewDialog(printer,parent)
    def render(p):
        painter=QPainter(p)
        pw=p.pageRect(QPrinter.Unit.DevicePixel).width()
        ph=p.pageRect(QPrinter.Unit.DevicePixel).height()
        y=80
        # Logo
        if S.logo_path and os.path.exists(S.logo_path):
            pix=QPixmap(S.logo_path).scaled(120,60,Qt.AspectRatioMode.KeepAspectRatio)
            painter.drawPixmap(40,20,pix)
            y=100
        painter.setFont(QFont("Segoe UI",14,QFont.Weight.Bold))
        painter.setPen(QColor(P['text'])); painter.drawText(40,y,title); y+=36
        painter.setFont(QFont("Segoe UI",9))
        painter.setPen(QColor(P['muted']))
        painter.drawText(40,y,f"Datum: {date.today().strftime('%d.%m.%Y.')}"); y+=36
        cw=int((pw-80)/max(len(headers),1))
        painter.fillRect(40,y-20,int(pw-80),30,QColor("#2E7
        # Dodati u main.py na GitHubu - PUT endpoint za korisnike
# Zamijeniti postojeci @app.delete("/korisnici/{kid}") sa ovim:
@app.put("/korisnici/{kid}")
def update_korisnik(kid: int, request: Request, user=Depends(get_current_user)):
    if user['uloga'] not in ('superadmin', 'admin'):
        raise HTTPException(403, "Nemate pristup")
    import json
    data = {}
    try:
        data = json.loads(request._body.decode()) if hasattr(request, '_body') else {}
    except:
        pass
    if 'ime' in data:
        db_exec("UPDATE korisnici SET ime=:i WHERE id=:k", {"i": data['ime'], "k": kid})
    if 'uloga' in data and user['uloga'] == 'superadmin':
        db_exec("UPDATE korisnici SET uloga=:u WHERE id=:k", {"u": data['uloga'], "k": kid})
    if 'nova_lozinka' in data and data['nova_lozinka']:
        import bcrypt
        pw = bcrypt.hashpw(data['nova_lozinka'].encode(), bcrypt.gensalt()).decode()
        db_exec("UPDATE korisnici SET password_hash=:p WHERE id=:k", {"p": pw, "k": kid})
    if 'aktivan' in data:
        db_exec("UPDATE korisnici SET aktivan=:a WHERE id=:k", {"a": 1 if data['aktivan'] else 0, "k": kid})
    return {"ok": True}

@app.delete("/korisnici/{kid}")
def del_korisnik(kid: int, request: Request, user=Depends(get_current_user)):
    if user['uloga'] not in ('superadmin', 'admin'):
        raise HTTPException(403, "Nemate pristup")
    db_exec("UPDATE korisnici SET aktivan=0 WHERE id=:k", {"k": kid})
    return {"ok": True}
